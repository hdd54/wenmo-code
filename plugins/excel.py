"""Excel 表格生成插件：用 openpyxl 生成 .xlsx（中文可靠，支持合并表头/样式/多工作表）。
模型不要手写 .xlsx（二进制格式会损坏/乱码）。"""

import os
import re
import tempfile
import time
from tenant_state import files_dir, resolve_scoped_file
from network_safety import safe_urlopen

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _safe(s, n=60):
    s = re.sub(r"[\\/:*?\"<>|]", "_", str(s or "表格"))
    return s[:n] or "表格"


def _resolve_image_local(src):
    """图片本地路径：data URL → 临时文件；http(s) → 下载临时文件；本地/files 路径 → 原路径"""
    import base64
    import tempfile
    import urllib.request
    s = str(src or "").strip()
    if not s:
        return None
    if s.startswith("data:image/"):
        m = re.match(r"^data:image/(\w+);base64,(.+)$", s, re.S)
        if not m:
            return None
        if len(m.group(2)) > 27_000_000:
            return None
        ext = {"png": "png", "jpeg": "jpg", "jpg": "jpg", "gif": "gif", "webp": "webp"}.get(m.group(1).lower(), "png")
        try:
            raw = base64.b64decode(m.group(2))
            if len(raw) > 20_000_000:
                return None
            fd, tmp = tempfile.mkstemp(suffix="." + ext)
            with os.fdopen(fd, "wb") as f:
                f.write(raw)
            return tmp
        except Exception:
            return None
    if s.startswith("http://") or s.startswith("https://"):
        try:
            req = urllib.request.Request(s, headers={"User-Agent": "Mozilla/5.0"})
            with safe_urlopen(req, timeout=15) as response:
                raw = response.read(20_000_001)
            if len(raw) > 20_000_000:
                return None
            fd, tmp = tempfile.mkstemp(suffix=".jpg")
            with os.fdopen(fd, "wb") as f:
                f.write(raw)
            return tmp
        except Exception:
            return None
    return resolve_scoped_file(s)


def create_xlsx(args):
    """用 openpyxl 生成 Excel 表格（支持多工作表、表头样式、列宽、合计行），保存到下载区并返回链接。"""
    title = str(args.get("title", "")).strip() or "表格"
    sheets = args.get("sheets") or []
    if not isinstance(sheets, list) or not sheets:
        return {"error": "sheets 参数不能为空：请提供 [{name, headers:[...], rows:[[...]]}] 列表"}
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return {"error": "服务器缺少 openpyxl，无法生成 Excel"}
    try:
        wb = Workbook()
        wb.remove(wb.active)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(style="thin", color="C0C0C0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for i, sh in enumerate(sheets):
            if not isinstance(sh, dict):
                continue
            name = str(sh.get("name", "")).strip() or f"Sheet{i + 1}"
            headers = sh.get("headers") or []
            rows = sh.get("rows") or []
            ws = wb.create_sheet(name[:31])
            # 表头
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=str(h))
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
            # 数据行
            for ri, row in enumerate(rows, 2):
                if not isinstance(row, (list, tuple)):
                    continue
                for ci, cell in enumerate(row, 1):
                    c = ws.cell(row=ri, column=ci, value=cell if cell is not None else "")
                    c.border = border
                    c.alignment = Alignment(vertical="center")
            # 列宽 + 冻结表头
            for col in ws.columns:
                letter = col[0].column_letter
                width = max(len(str(c.value or "")) for c in col[:50]) if col else 10
                ws.column_dimensions[letter].width = min(max(width * 1.4 + 4, 10), 40)
            if headers:
                ws.freeze_panes = "A2"
            # 图片：sheet 数据区下方插入（src 支持 http(s) URL / data URL / 本地路径）
            images = sh.get("images") or sh.get("image") or []
            if isinstance(images, dict):
                images = [images]
            if isinstance(images, list):
                from openpyxl.drawing.image import Image as XLImage
                anchor_row = len(rows) + 3   # 数据下方空两行
                for img_i, img in enumerate(images):
                    if not isinstance(img, dict):
                        continue
                    src = str(img.get("src") or img.get("image") or img.get("url") or "").strip()
                    if not src:
                        continue
                    img_path = _resolve_image_local(src)
                    if not img_path:
                        continue
                    try:
                        xl = XLImage(img_path)
                        # 指定 width（英寸语义，转像素）则等比缩放；不指定保持原尺寸
                        w_in = float(img.get("width") or 0)
                        if w_in > 0 and xl.width:
                            scale = min(w_in * 96 / xl.width, 1.0)
                            xl.width = int(xl.width * scale)
                            if xl.height:
                                xl.height = int(xl.height * scale)
                        ws.add_image(xl, f"A{anchor_row + img_i * 18}")
                    except Exception:
                        pass
                    finally:
                        if img_path.startswith(tempfile.gettempdir()):
                            try:
                                os.remove(img_path)
                            except Exception:
                                pass
        fname = f"{_safe(title)}_{int(time.time())}.xlsx"
        fpath = os.path.join(files_dir(), fname)
        wb.save(fpath)
        url = f"http://127.0.0.1:8000/files/{fname}"
        return {
            "ok": True,
            "file": fname,
            "url": url,
            "note": f"Excel 已生成：{fname}（{len(sheets)} 个工作表，带表头样式/列宽/冻结首行）。请在回答里引用链接。",
        }
    except Exception as e:
        return {"error": f"Excel 生成失败: {e}"}


PLUGIN_TOOLS = [
    {
        "name": "create_xlsx",
        "description": "创建 Excel 表格（服务器 openpyxl 生成，中文可靠；支持多工作表、表头样式、自动列宽、冻结首行、图片嵌入，排版规范）。"
                       "当用户要求『做一个 Excel/表格/数据表/统计表/名单』时，必须使用本工具，不要自己手写 .xlsx。"
                       "参数：title=表格名；sheets=[{name: 工作表名, headers: [列名], rows: [[单元格值]], "
                       "images: [{src: 图片URL/base64/路径, width: 宽英寸(可选)}]}]（images 可选，用于表格配图）。"
                       "生成后返回链接，请在回答中引用。",
        "parameters": {
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "表格名称"},
                "sheets": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string", "description": "工作表名"},
                            "headers": {"type": "array", "items": {"type": "string"}, "description": "列名列表"},
                            "rows": {"type": "array", "items": {"type": "array"}, "description": "数据行列表"}
                        },
                        "required": ["name", "headers", "rows"]
                    },
                    "description": "工作表列表（最多 10 个）"
                }
            },
            "required": ["title", "sheets"],
        },
        "handler": create_xlsx,
    }
]
